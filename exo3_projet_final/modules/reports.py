"""
Module de génération de rapports
"""

import json
import csv
import os
import logging
from datetime import datetime, timedelta
from typing import Dict, Any, List
import pandas as pd
import matplotlib.pyplot as plt
import openpyxl
from openpyxl.styles import Font, PatternFill

logger = logging.getLogger(__name__)


class ReportGenerator:
    """Classe pour la génération de rapports réseau"""
    
    def __init__(self, report_dir: str = 'reports'):
        self.report_dir = report_dir
        os.makedirs(report_dir, exist_ok=True)
        logger.info(f"ReportGenerator initialisé avec report_dir={report_dir}")
    
    def parse_uptime(self, uptime_str: str) -> timedelta:
        """Parser la chaîne uptime en objet timedelta"""
        try:
            if isinstance(uptime_str, (int, float)):
                return timedelta(seconds=uptime_str)
            
            # Formats communs: "1 week, 2 days, 3 hours, 4 minutes", "5 days, 6:30:00", etc.
            total_seconds = 0
            
            # Séparer par virgules
            parts = str(uptime_str).split(',')
            for part in parts:
                part = part.strip().lower()
                
                if 'week' in part:
                    weeks = int(part.split()[0])
                    total_seconds += weeks * 7 * 24 * 3600
                elif 'day' in part:
                    days = int(part.split()[0])
                    total_seconds += days * 24 * 3600
                elif 'hour' in part:
                    hours = int(part.split()[0])
                    total_seconds += hours * 3600
                elif 'minute' in part:
                    minutes = int(part.split()[0])
                    total_seconds += minutes * 60
                elif 'second' in part:
                    seconds = int(part.split()[0])
                    total_seconds += seconds
                elif ':' in part:  # Format HH:MM:SS
                    time_parts = part.split(':')
                    if len(time_parts) == 3:
                        hours, minutes, seconds = map(int, time_parts)
                        total_seconds += hours * 3600 + minutes * 60 + seconds
            
            return timedelta(seconds=total_seconds)
            
        except Exception as e:
            logger.warning(f"Impossible de parser uptime '{uptime_str}': {e}")
            return timedelta(0)
    
    def generate_comprehensive_report(self, devices: Dict[str, Any], 
                                   monitoring_data: List[Dict] = None,
                                   export_format: str = 'all') -> Dict[str, str]:
        """Générer un rapport complet avec analyse avancée"""
        logger.info("Génération rapport complet...")
        
        from modules.napalm_utils import NetworkAutomation
        auto = NetworkAutomation()
        
        report_data = {
            'metadata': {
                'generated_at': datetime.now().isoformat(),
                'devices_analyzed': len(devices),
                'report_version': '2.0'
            },
            'summary': {},
            'device_details': {},
            'monitoring_stats': {}
        }
        
        # Collecte des données
        for device_name, device_config in devices.items():
            logger.info(f"Analyse de {device_name}...")
            device_info = auto.collect_device_info(device_config)
            report_data['device_details'][device_name] = device_info
        
        # Statistiques de monitoring si disponibles
        if monitoring_data:
            from modules.monitoring import NetworkMonitoring
            monitor = NetworkMonitoring()
            monitor.monitoring_data = monitoring_data
            
            report_data['monitoring_stats'] = {
                device_name: monitor.calculate_uptime_stats(device_name)
                for device_name in devices.keys()
            }
        
        # Génération des exports
        report_files = {}
        
        if export_format in ['json', 'all']:
            json_report = self._save_json_report(report_data)
            report_files['json'] = json_report
        
        if export_format in ['csv', 'all']:
            csv_report = self.generate_csv_report(report_data)
            report_files['csv'] = csv_report
        
        if export_format in ['excel', 'all']:
            excel_report = self.generate_excel_report(report_data)
            report_files['excel'] = excel_report
        
        # Génération des graphiques
        charts = self.generate_advanced_charts(report_data)
        if charts:
            report_files['charts'] = charts
        
        logger.info(f"Rapport complet généré: {len(report_files)} fichier(s)")
        return report_files
    
    def _save_json_report(self, report_data: Dict[str, Any]) -> str:
        """Sauvegarder le rapport JSON"""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"{self.report_dir}/network_report_{timestamp}.json"
        
        with open(filename, 'w', encoding='utf-8') as f:
            json.dump(report_data, f, indent=2, ensure_ascii=False)
        
        return filename
    
    def generate_csv_report(self, report_data: Dict[str, Any]) -> str:
        """Générer un rapport CSV détaillé"""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"{self.report_dir}/device_summary_{timestamp}.csv"
        
        with open(filename, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            # En-tête étendue
            writer.writerow([
                'Device', 'Hostname', 'Model', 'OS Version', 
                'Uptime (jours)', 'Uptime (secondes)', 'Interface Count',
                'CPU Usage (%)', 'Memory Usage (%)', 'Last Backup'
            ])
            
            for device_name, details in report_data['device_details'].items():
                if 'facts' in details:
                    facts = details['facts']
                    
                    # Calcul uptime
                    uptime_td = self.parse_uptime(facts.get('uptime', 0))
                    uptime_days = uptime_td.total_seconds() / (24 * 3600)
                    uptime_seconds = uptime_td.total_seconds()
                    
                    # Usage CPU/Mémoire
                    environment = details.get('environment', {})
                    cpu = environment.get('cpu', 0)
                    memory = environment.get('memory', {})
                    memory_usage = memory.get('used_ram', 0)
                    
                    writer.writerow([
                        device_name,
                        facts.get('hostname', 'N/A'),
                        facts.get('model', 'N/A'),
                        facts.get('os_version', 'N/A'),
                        f"{uptime_days:.2f}",
                        int(uptime_seconds),
                        len(details.get('interfaces', {})),
                        cpu,
                        memory_usage,
                        'Yes' if details.get('running_config') else 'No'
                    ])
        
        logger.info(f"Rapport CSV généré: {filename}")
        return filename
    
    def generate_excel_report(self, report_data: Dict[str, Any]) -> str:
        """Générer un rapport Excel avec mise en forme"""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"{self.report_dir}/network_report_{timestamp}.xlsx"
        
        wb = openpyxl.Workbook()
        
        # Feuille de résumé
        ws_summary = wb.active
        ws_summary.title = "Résumé"
        
        # En-têtes avec style
        headers = ['Device', 'Hostname', 'Modèle', 'OS', 'Uptime (jours)', 'Interfaces', 'Statut']
        for col, header in enumerate(headers, 1):
            cell = ws_summary.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        
        # Données
        row = 2
        for device_name, details in report_data['device_details'].items():
            if 'facts' in details:
                facts = details['facts']
                uptime_td = self.parse_uptime(facts.get('uptime', 0))
                uptime_days = uptime_td.total_seconds() / (24 * 3600)
                
                ws_summary.cell(row=row, column=1, value=device_name)
                ws_summary.cell(row=row, column=2, value=facts.get('hostname', 'N/A'))
                ws_summary.cell(row=row, column=3, value=facts.get('model', 'N/A'))
                ws_summary.cell(row=row, column=4, value=facts.get('os_version', 'N/A'))
                ws_summary.cell(row=row, column=5, value=uptime_days)
                ws_summary.cell(row=row, column=6, value=len(details.get('interfaces', {})))
                ws_summary.cell(row=row, column=7, value='OK')
                
                row += 1
        
        # Feuille monitoring si disponible
        if report_data.get('monitoring_stats'):
            ws_monitoring = wb.create_sheet("Monitoring")
            monitoring_headers = ['Device', 'Disponibilité (%)', 'Temps réponse moyen (s)', 'Dernier statut']
            
            for col, header in enumerate(monitoring_headers, 1):
                cell = ws_monitoring.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            
            row = 2
            for device_name, stats in report_data['monitoring_stats'].items():
                if not isinstance(stats, dict) or 'error' in stats:
                    continue
                
                ws_monitoring.cell(row=row, column=1, value=device_name)
                ws_monitoring.cell(row=row, column=2, value=stats.get('availability_percent', 0))
                ws_monitoring.cell(row=row, column=3, value=stats.get('avg_response_time', 0))
                ws_monitoring.cell(row=row, column=4, value=stats.get('last_status', 'UNKNOWN'))
                row += 1
        
        wb.save(filename)
        logger.info(f"Rapport Excel généré: {filename}")
        return filename
    
    def generate_advanced_charts(self, report_data: Dict[str, Any]) -> List[str]:
        """Générer des graphiques avancés"""
        try:
            devices = []
            interface_counts = []
            uptime_days = []
            availability_pct = []
            
            for device_name, details in report_data['device_details'].items():
                if 'facts' in details:
                    facts = details['facts']
                    devices.append(device_name)
                    interface_counts.append(len(details.get('interfaces', {})))
                    
                    # Uptime réel
                    uptime_td = self.parse_uptime(facts.get('uptime', 0))
                    uptime_days.append(uptime_td.total_seconds() / (24 * 3600))
            
            # Statistiques de monitoring
            monitoring_stats = report_data.get('monitoring_stats', {})
            for device_name in devices:
                stats = monitoring_stats.get(device_name, {})
                availability_pct.append(stats.get('availability_percent', 100))
            
            # Graphique 1: Interfaces
            plt.figure(figsize=(12, 6))
            plt.subplot(1, 2, 1)
            bars = plt.bar(devices, interface_counts, color='skyblue', alpha=0.7)
            plt.title('Interfaces par équipement')
            plt.xlabel('Équipements')
            plt.ylabel('Nombre d interfaces')
            plt.xticks(rotation=45)
            
            # Ajouter les valeurs sur les barres
            for bar in bars:
                height = bar.get_height()
                plt.text(bar.get_x() + bar.get_width()/2., height,
                        f'{int(height)}', ha='center', va='bottom')
            
            # Graphique 2: Uptime
            plt.subplot(1, 2, 2)
            bars = plt.bar(devices, uptime_days, color='lightgreen', alpha=0.7)
            plt.title('Uptime des équipements (jours)')
            plt.xlabel('Équipements')
            plt.ylabel('Uptime (jours)')
            plt.xticks(rotation=45)
            
            for bar in bars:
                height = bar.get_height()
                plt.text(bar.get_x() + bar.get_width()/2., height,
                        f'{height:.1f}', ha='center', va='bottom')
            
            plt.tight_layout()
            chart1 = f'{self.report_dir}/equipment_charts.png'
            plt.savefig(chart1, dpi=150, bbox_inches='tight')
            plt.close()
            
            # Graphique 3: Disponibilité (si données monitoring)
            if availability_pct and any(pct != 100 for pct in availability_pct):
                plt.figure(figsize=(10, 6))
                colors = ['green' if pct > 95 else 'orange' if pct > 90 else 'red' 
                         for pct in availability_pct]
                
                bars = plt.bar(devices, availability_pct, color=colors, alpha=0.7)
                plt.title('Disponibilité des équipements (%)')
                plt.xlabel('Équipements')
                plt.ylabel('Disponibilité (%)')
                plt.xticks(rotation=45)
                plt.ylim(0, 100)
                
                for bar in bars:
                    height = bar.get_height()
                    plt.text(bar.get_x() + bar.get_width()/2., height,
                            f'{height:.1f}%', ha='center', va='bottom')
                
                plt.tight_layout()
                chart2 = f'{self.report_dir}/availability_chart.png'
                plt.savefig(chart2, dpi=150, bbox_inches='tight')
                plt.close()
                
                return [chart1, chart2]
            
            return [chart1]
            
        except Exception as e:
            logger.error(f"Erreur génération graphiques: {e}")
            return []